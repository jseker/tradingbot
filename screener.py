import json
import smtplib
import ssl
from datetime import datetime, date, timedelta
import pandas as pd
import numpy as np
import yfinance as yf
import pandas_market_calendars as mcal
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

POSITIONS_FILE = 'C:\\TradingBot\\positions.xlsx'
POSITIONS_TEMP = 'C:\\TradingBot\\positions_temp.xlsx'

def safe_save(wb):
    import shutil, os
    try:
        wb.save(POSITIONS_TEMP)
        test = openpyxl.load_workbook(POSITIONS_TEMP)
        test.close()
        os.replace(POSITIONS_TEMP, POSITIONS_FILE)
        return True
    except PermissionError:
        print('  WARNING: positions.xlsx is open in Excel — skipping update to avoid corruption.')
        print('  Close Excel before running the screener for TechnicalSnapshot updates.')
        try:
            if os.path.exists(POSITIONS_TEMP): os.remove(POSITIONS_TEMP)
        except: pass
        return False
    except Exception as ex:
        print('  WARNING: Could not save positions.xlsx: ' + str(ex))
        try:
            if os.path.exists(POSITIONS_TEMP): os.remove(POSITIONS_TEMP)
        except: pass
        return False

# ── STYLE HELPERS ────────────────────────────────────────────────────
def bg(h): return PatternFill(start_color=h, end_color=h, fill_type='solid')

# ── MACRO CALENDAR ───────────────────────────────────────────────────
def get_nth_weekday(year, month, n, weekday):
    first = date(year, month, 1)
    days_until = (weekday - first.weekday()) % 7
    return first + timedelta(days=days_until) + timedelta(weeks=n-1)

def get_last_weekday(year, month, weekday):
    last = date(year+1, 1, 1) - timedelta(days=1) if month == 12 else date(year, month+1, 1) - timedelta(days=1)
    return last - timedelta(days=(last.weekday() - weekday) % 7)

def get_first_business_day(year, month):
    d = date(year, month, 1)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d

def get_macro_events(start_date, end_date):
    events = []
    fomc_2026 = [date(2026,1,28),date(2026,3,18),date(2026,5,6),date(2026,6,17),
                 date(2026,7,29),date(2026,9,16),date(2026,11,4),date(2026,12,16)]
    current = start_date.replace(day=1)
    while current <= end_date + timedelta(days=31):
        y, m = current.year, current.month
        cpi = get_nth_weekday(y, m, 2, 2)
        if start_date <= cpi <= end_date:
            events.append({'date':cpi,'name':'CPI','impact':'HIGH','note':'Major inflation indicator — can move tech stocks 2-4%'})
        ppi = cpi + timedelta(days=1)
        if start_date <= ppi <= end_date:
            events.append({'date':ppi,'name':'PPI','impact':'MEDIUM','note':'Leading indicator for future CPI'})
        nfp = get_nth_weekday(y, m, 1, 4)
        if start_date <= nfp <= end_date:
            events.append({'date':nfp,'name':'Non-Farm Payrolls','impact':'HIGH','note':'Jobs report — strong jobs = rate fears'})
        for fd in fomc_2026:
            if start_date <= fd <= end_date:
                events.append({'date':fd,'name':'FOMC Rate Decision','impact':'HIGH','note':'Fed rate decision — major market mover'})
        pce = get_last_weekday(y, m, 4)
        if start_date <= pce <= end_date:
            events.append({'date':pce,'name':'PCE Inflation','impact':'HIGH','note':"Fed's preferred inflation measure"})
        retail = get_nth_weekday(y, m, 2, 1)
        if start_date <= retail <= end_date:
            events.append({'date':retail,'name':'Retail Sales','impact':'MEDIUM','note':'Consumer spending indicator'})
        ism = get_first_business_day(y, m)
        if start_date <= ism <= end_date:
            events.append({'date':ism,'name':'ISM Manufacturing PMI','impact':'MEDIUM','note':'Economic activity indicator'})
        conf = get_last_weekday(y, m, 1)
        if start_date <= conf <= end_date:
            events.append({'date':conf,'name':'Consumer Confidence','impact':'LOW','note':'Consumer sentiment survey'})
        thu = get_nth_weekday(y, m, 1, 3)
        while thu.month == m:
            if start_date <= thu <= end_date and (thu - start_date).days <= 10:
                events.append({'date':thu,'name':'Weekly Jobless Claims','impact':'LOW','note':'Weekly labor market check'})
            thu += timedelta(days=7)
        current = current.replace(year=y+1, month=1) if m == 12 else current.replace(month=m+1)
    seen = set()
    result = []
    for e in sorted(events, key=lambda x: x['date']):
        k = (e['date'], e['name'])
        if k not in seen:
            seen.add(k)
            result.append(e)
    return result

def get_macro_flag_string(expiry_date, macro_events):
    if not expiry_date:
        return ''
    today = date.today()
    if hasattr(expiry_date, 'date'):
        expiry_date = expiry_date.date()
    relevant = [e for e in macro_events if e['impact'] in ('HIGH','MEDIUM') and today <= e['date'] <= expiry_date]
    if not relevant:
        return 'None'
    parts = []
    for e in relevant:
        parts.append(e['name'] + ' ' + e['date'].strftime('%b %d'))
    return ', '.join(parts)

# ── MARKET DATA ──────────────────────────────────────────────────────
def get_vix():
    try:
        vix = yf.Ticker('^VIX')
        info = vix.info
        val = info.get('regularMarketPrice') or info.get('currentPrice')
        if val:
            return round(float(val), 2)
        hist = vix.history(period='1d')
        return round(float(hist['Close'].iloc[-1]), 2) if not hist.empty else None
    except:
        return None

def get_spy_trend():
    try:
        spy = yf.Ticker('SPY')
        # Try 6mo first for more data, fall back to 3mo
        hist = spy.history(period='6mo')
        if hist is None or hist.empty:
            hist = spy.history(period='3mo')
        if hist is None or hist.empty:
            return None
        # Drop any NaN rows entirely
        hist = hist.dropna(subset=['Close'])
        closes = hist['Close'].astype(float)
        if len(closes) < 50:
            return None
        # Use last valid close — sometimes today's bar has NaN if pre-market
        # Walk back until we find a valid non-NaN close
        current = None
        for i in range(len(closes)-1, -1, -1):
            val = closes.iloc[i]
            if not np.isnan(val) and val > 0:
                current = float(val)
                break
        if current is None:
            return None
        # Previous closes for change calculation
        valid_closes = closes[~np.isnan(closes) & (closes > 0)]
        prev1 = float(valid_closes.iloc[-2]) if len(valid_closes) >= 2 else current
        prev5 = float(valid_closes.iloc[-6]) if len(valid_closes) >= 6 else current
        sma20 = float(valid_closes.rolling(20).mean().dropna().iloc[-1])
        sma50 = float(valid_closes.rolling(50).mean().dropna().iloc[-1])
        if np.isnan(sma20) or np.isnan(sma50):
            return None
        change_1d = round((current - prev1) / prev1 * 100, 2) if prev1 > 0 else 0
        change_5d = round((current - prev5) / prev5 * 100, 2) if prev5 > 0 else 0
        if current > sma20 and current > sma50:
            trend = 'UPTREND'
        elif current < sma20 and current < sma50:
            trend = 'DOWNTREND'
        else:
            trend = 'MIXED'
        return {'current':round(current,2),'sma20':round(sma20,2),'sma50':round(sma50,2),
                'trend':trend,'change_1d':change_1d,'change_5d':change_5d}
    except Exception as ex:
        print('Warning: SPY trend fetch failed: ' + str(ex))
        return None

def get_market_regime(vix, spy_trend):
    if vix is None or spy_trend is None:
        return 'UNKNOWN', ['Could not fetch market data']
    verdict = 'FAVORABLE'
    reasons = []
    if spy_trend['trend'] == 'UPTREND':
        reasons.append('SPY in uptrend — good backdrop for put writing')
    elif spy_trend['trend'] == 'DOWNTREND':
        verdict = 'UNFAVORABLE'
        reasons.append('SPY in downtrend — elevated assignment risk')
    else:
        verdict = 'NEUTRAL'
        reasons.append('SPY trend mixed — proceed with caution')
    if vix < 15:
        if verdict == 'FAVORABLE': verdict = 'NEUTRAL'
        reasons.append('VIX ' + str(vix) + ' — low IV, premiums thin')
    elif vix < 20:
        reasons.append('VIX ' + str(vix) + ' — moderate IV, normal premium environment')
    elif vix < 30:
        reasons.append('VIX ' + str(vix) + ' — elevated IV, excellent premium environment')
    else:
        if verdict != 'UNFAVORABLE': verdict = 'NEUTRAL'
        reasons.append('VIX ' + str(vix) + ' — extreme fear, rich premiums but high assignment risk')
    return verdict, reasons

# ── STOCK ANALYSIS ───────────────────────────────────────────────────
def get_analyst_data(ticker):
    try:
        info = yf.Ticker(ticker).info
        target = info.get('targetMeanPrice')
        current = info.get('currentPrice') or info.get('regularMarketPrice')
        rec = info.get('recommendationMean')
        n_analysts = info.get('numberOfAnalystOpinions')
        beta = info.get('beta')
        fwd_pe = info.get('forwardPE')
        upside = round((target - current) / current * 100, 1) if target and current and current > 0 else None
        rec_map = {1.5:'STRONG BUY', 2.5:'BUY', 3.5:'HOLD', 4.5:'UNDERPERFORM'}
        rec_text = None
        if rec:
            for threshold, label in sorted(rec_map.items()):
                if rec <= threshold:
                    rec_text = label
                    break
            if rec_text is None:
                rec_text = 'SELL'
        beta_text = None
        beta_risk = 'MODERATE'
        if beta:
            beta = round(beta, 2)
            if beta > 1.8: beta_text, beta_risk = str(beta) + ' — VERY HIGH BETA', 'VERY HIGH'
            elif beta > 1.3: beta_text, beta_risk = str(beta) + ' — HIGH BETA', 'HIGH'
            elif beta > 0.8: beta_text, beta_risk = str(beta) + ' — MODERATE BETA', 'MODERATE'
            else: beta_text, beta_risk = str(beta) + ' — LOW BETA (defensive)', 'LOW'
        return {'target':round(target,2) if target else None,'current':round(current,2) if current else None,
                'upside':upside,'rec':rec_text,'n_analysts':n_analysts,'beta':beta,
                'beta_text':beta_text,'beta_risk':beta_risk,'fwd_pe':round(fwd_pe,1) if fwd_pe else None}
    except:
        return None

def get_iv_data(ticker):
    try:
        stock = yf.Ticker(ticker)
        hist = stock.history(period='1y')
        if len(hist) < 20:
            return None
        closes = hist['Close']
        returns = closes.pct_change().dropna()
        hist_vol = float(returns.rolling(20).std().iloc[-1]) * np.sqrt(252) * 100
        opt = stock.options
        if not opt:
            return {'iv_rank':None,'hist_vol':round(hist_vol,1),'iv_estimate':None,'iv_signal':'NEUTRAL'}
        chain = stock.option_chain(opt[0])
        current_price = float(closes.iloc[-1])
        atm = chain.puts[abs(chain.puts['strike'] - current_price) < current_price * 0.05]
        if atm.empty:
            return {'iv_rank':None,'hist_vol':round(hist_vol,1),'iv_estimate':None,'iv_signal':'NEUTRAL'}
        iv_est = float(atm['impliedVolatility'].mean()) * 100
        ratio = iv_est / hist_vol * 100 if hist_vol > 0 else 100
        if ratio > 120:
            iv_signal, iv_text = 'GOOD', 'HIGH (' + str(round(iv_est,1)) + '% IV vs ' + str(round(hist_vol,1)) + '% HV) — good time to sell premium'
        elif ratio > 90:
            iv_signal, iv_text = 'NEUTRAL', 'NORMAL (' + str(round(iv_est,1)) + '% IV vs ' + str(round(hist_vol,1)) + '% HV)'
        else:
            iv_signal, iv_text = 'POOR', 'LOW (' + str(round(iv_est,1)) + '% IV vs ' + str(round(hist_vol,1)) + '% HV) — premiums compressed'
        return {'iv_rank':iv_text,'iv_signal':iv_signal,'hist_vol':round(hist_vol,1),'iv_estimate':round(iv_est,1)}
    except:
        return None

def get_earnings_date(ticker):
    try:
        cal = yf.Ticker(ticker).calendar
        if cal is not None and not cal.empty:
            earn_date = pd.Timestamp(cal.iloc[0, 0])
            days_away = (earn_date - pd.Timestamp.now()).days
            return earn_date.strftime('%Y-%m-%d'), days_away
        return None, None
    except:
        return None, None

def calculate_technicals(ticker):
    try:
        hist = yf.Ticker(ticker).history(period='3mo')
        if hist is None or hist.empty or len(hist) < 20:
            return None
        closes = hist['Close'].dropna()
        volumes = hist['Volume'].dropna()
        if len(closes) < 20:
            return None
        current = float(closes.iloc[-1])
        if np.isnan(current) or current <= 0:
            return None
        delta = closes.diff()
        gain = delta.where(delta > 0, 0)
        loss = -delta.where(delta < 0, 0)
        avg_gain = gain.rolling(14).mean()
        avg_loss = loss.rolling(14).mean()
        rs = avg_gain / avg_loss.replace(0, np.nan)
        rsi_series = 100 - (100 / (1 + rs))
        rsi_val = float(rsi_series.dropna().iloc[-1]) if not rsi_series.dropna().empty else 50.0
        if np.isnan(rsi_val):
            rsi_val = 50.0
        sma20 = closes.rolling(20).mean()
        std20 = closes.rolling(20).std()
        upper_bb = sma20 + 2*std20
        lower_bb = sma20 - 2*std20
        sma20v = float(sma20.dropna().iloc[-1]) if not sma20.dropna().empty else current
        lower_v = float(lower_bb.dropna().iloc[-1]) if not lower_bb.dropna().empty else current * 0.95
        upper_v = float(upper_bb.dropna().iloc[-1]) if not upper_bb.dropna().empty else current * 1.05
        sma50_series = closes.rolling(50).mean().dropna()
        sma50v = float(sma50_series.iloc[-1]) if len(closes) >= 50 and not sma50_series.empty else None
        bb_pos = round((current - lower_v) / (upper_v - lower_v), 2) if (upper_v - lower_v) > 0 else 0.5
        pct_sma20 = round((current - sma20v) / sma20v * 100, 2) if sma20v > 0 else 0
        pct_sma50 = round((current - sma50v) / sma50v * 100, 2) if sma50v and sma50v > 0 else None
        avg_vol = float(volumes.rolling(20).mean().dropna().iloc[-1]) if not volumes.rolling(20).mean().dropna().empty else 1
        vol_ratio = round(float(volumes.iloc[-1]) / avg_vol, 2) if avg_vol > 0 else 1.0
        atr_series = (hist['High'] - hist['Low']).rolling(14).mean().dropna()
        atr = round(float(atr_series.iloc[-1]), 2) if not atr_series.empty else 0
        atr_pct = round(atr / current * 100, 2) if current > 0 else 0
        prev1 = float(closes.iloc[-2]) if len(closes) >= 2 else current
        prev5 = float(closes.iloc[-6]) if len(closes) >= 6 else current
        chg1d = round((current - prev1) / prev1 * 100, 2) if prev1 > 0 else 0
        chg5d = round((current - prev5) / prev5 * 100, 2) if prev5 > 0 else 0
        return {
            'ticker': ticker, 'current': round(current, 2), 'rsi': round(rsi_val, 1),
            'sma20': round(sma20v, 2), 'sma50': round(sma50v, 2) if sma50v else None,
            'lower_band': round(lower_v, 2), 'upper_band': round(upper_v, 2),
            'bb_position': bb_pos, 'pct_from_sma20': pct_sma20, 'pct_from_sma50': pct_sma50,
            'volume_ratio': vol_ratio, 'atr': atr, 'atr_pct': atr_pct,
            'change_1d': chg1d, 'change_5d': chg5d
        }
    except Exception as ex:
        print('  Technicals failed for ' + ticker + ': ' + str(ex))
        return None

def generate_recommendation(ticker, tech, analyst, iv_data, earnings_date, earnings_days, macro_warn, cfg):
    if tech is None:
        return 'INSUFFICIENT DATA', [], [], 0
    reasons_for, reasons_against, score = [], [], 0
    if tech['rsi'] < 25:
        score += 3; reasons_for.append('RSI ' + str(tech['rsi']) + ' — extremely oversold')
    elif tech['rsi'] < 30:
        score += 2; reasons_for.append('RSI ' + str(tech['rsi']) + ' — oversold')
    elif tech['rsi'] < 35:
        score += 1; reasons_for.append('RSI ' + str(tech['rsi']) + ' — approaching oversold')
    elif tech['rsi'] > 65:
        score -= 2; reasons_against.append('RSI ' + str(tech['rsi']) + ' — elevated, not oversold')
    if tech['bb_position'] < 0.05:
        score += 3; reasons_for.append('At/below lower Bollinger Band — statistically extreme')
    elif tech['bb_position'] < 0.20:
        score += 2; reasons_for.append('Near lower Bollinger Band (position: ' + str(tech['bb_position']) + ')')
    elif tech['bb_position'] < 0.30:
        score += 1; reasons_for.append('Below lower quarter of Bollinger Bands')
    if tech['pct_from_sma20'] < -10:
        score += 2; reasons_for.append(str(abs(tech['pct_from_sma20'])) + '% below 20d SMA — extended move')
    elif tech['pct_from_sma20'] < -5:
        score += 1; reasons_for.append(str(abs(tech['pct_from_sma20'])) + '% below 20d SMA')
    if tech['volume_ratio'] < 0.7:
        score += 1; reasons_for.append('Low volume (' + str(tech['volume_ratio']) + 'x avg) — likely technical not fundamental')
    elif tech['volume_ratio'] > 2.5:
        score -= 1; reasons_against.append('High volume (' + str(tech['volume_ratio']) + 'x avg) — possible fundamental selling')
    if analyst:
        if analyst['upside'] and analyst['upside'] > 20:
            score += 2; reasons_for.append('Analyst target $' + str(analyst['target']) + ' — ' + str(analyst['upside']) + '% upside')
        elif analyst['upside'] and analyst['upside'] > 10:
            score += 1; reasons_for.append('Analyst target $' + str(analyst['target']) + ' — ' + str(analyst['upside']) + '% upside')
        elif analyst['upside'] and analyst['upside'] < 0:
            score -= 2; reasons_against.append('Analyst target BELOW current price — avoid')
        if analyst['rec'] in ['STRONG BUY', 'BUY']:
            score += 1; reasons_for.append('Analyst consensus: ' + analyst['rec'] + ' (' + str(analyst['n_analysts']) + ' analysts)')
        elif analyst['rec'] in ['UNDERPERFORM', 'SELL']:
            score -= 2; reasons_against.append('Analyst consensus: ' + analyst['rec'])
        if analyst['beta_risk'] == 'VERY HIGH':
            score -= 1; reasons_against.append('Beta ' + str(analyst['beta']) + ' — very high beta, size conservatively')
        elif analyst['beta_risk'] == 'HIGH':
            reasons_against.append('Beta ' + str(analyst['beta']) + ' — high beta, size conservatively')
    if iv_data:
        if iv_data.get('iv_signal') == 'GOOD':
            score += 2; reasons_for.append('IV elevated vs historical — premium environment favorable')
        elif iv_data.get('iv_signal') == 'POOR':
            score -= 1; reasons_against.append('IV compressed — premiums thin, consider waiting')
    if earnings_days is not None:
        if earnings_days <= 5:
            score -= 5; reasons_against.append('EARNINGS IN ' + str(earnings_days) + ' DAYS — DO NOT write puts')
        elif earnings_days <= 14:
            score -= 2; reasons_against.append('Earnings in ' + str(earnings_days) + ' days — elevated event risk')
        elif earnings_days <= 30:
            score -= 1; reasons_against.append('Earnings in ' + str(earnings_days) + ' days — consider shorter expiry')
    if macro_warn:
        score -= 1; reasons_against.append('High impact macro event within 3 days: ' + macro_warn)
    if score <= -3 or (earnings_days is not None and earnings_days <= 5):
        verdict = 'AVOID'
    elif score <= 0:
        verdict = 'WAIT — conditions not favorable'
    elif score <= 3:
        verdict = 'WATCH — some signals present'
    elif score <= 5:
        verdict = 'CONSIDER — moderate conviction'
    else:
        verdict = 'WRITE PUT — high conviction'
    return verdict, reasons_for, reasons_against, score

# ── EXCEL AUTOMATION ─────────────────────────────────────────────────
def update_excel(open_puts, macro_events, tech_data_cache):
    try:
        wb = openpyxl.load_workbook(POSITIONS_FILE)

        # ── 1. Update MacroRiskFlag in OpenPuts (col 14) ─────────────
        ws_puts = wb['OpenPuts']
        puts_headers = [str(c.value).replace('\n','').replace(' ','') if c.value else None for c in ws_puts[3]]
        expiry_col = None
        flag_col = None
        for i, h in enumerate(puts_headers):
            if h == 'Expiry': expiry_col = i + 1
            if h == 'MacroRiskFlag': flag_col = i + 1
        if expiry_col and flag_col:
            for row in range(4, ws_puts.max_row + 1):
                ticker_val = ws_puts.cell(row=row, column=1).value
                if ticker_val is None:
                    continue
                expiry_val = ws_puts.cell(row=row, column=expiry_col).value
                if expiry_val:
                    flag = get_macro_flag_string(expiry_val, macro_events)
                    c = ws_puts.cell(row=row, column=flag_col)
                    c.value = flag
                    c.fill = bg('152236')
                    c.font = Font(color='FFB347' if flag and flag != 'None' else '00E5CC',
                                  size=9, name='Calibri')
                    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            print('  MacroRiskFlag updated in OpenPuts')

        # ── 2. Update TechnicalSnapshot sheet ────────────────────────
        ws_ts = wb['TechnicalSnapshot']
        ws_ts['A3'] = '  Last Updated: ' + str(date.today())
        ws_ts['A3'].fill = bg('1A3A5C')
        ws_ts['A3'].font = Font(color='00B4D8', bold=True, size=10, name='Calibri')

        # Find which tickers are in the sheet
        ticker_rows = {}
        for row in range(5, ws_ts.max_row + 1):
            val = ws_ts.cell(row=row, column=1).value
            if val and str(val).strip():
                ticker_rows[str(val).strip()] = row

        # Add any new tickers from watchlist not yet in sheet
        next_row = max(ticker_rows.values()) + 1 if ticker_rows else 5
        all_tickers_in_cache = list(tech_data_cache.keys())
        for ticker in all_tickers_in_cache:
            if ticker not in ticker_rows:
                ticker_rows[ticker] = next_row
                next_row += 1

        # Write technical data for each ticker
        for ticker, row in sorted(ticker_rows.items(), key=lambda x: x[0]):
            tech = tech_data_cache.get(ticker)
            # Style all cells in row
            for col in range(1, 12):
                c = ws_ts.cell(row=row, column=col)
                c.fill = bg('152236')
                c.font = Font(color='FFFFFF', size=10, name='Calibri')
                c.alignment = Alignment(horizontal='center', vertical='center')
                from openpyxl.styles import Border, Side
                s = Side(style='thin', color='1E3A5F')
                c.border = Border(left=s, right=s, top=s, bottom=s)
            ws_ts.cell(row=row, column=1).value = ticker
            ws_ts.cell(row=row, column=1).font = Font(color='00E5CC', bold=True, size=11, name='Calibri')
            ws_ts.row_dimensions[row].height = 20
            if tech is None:
                ws_ts.cell(row=row, column=11).value = 'No data'
                ws_ts.cell(row=row, column=11).font = Font(color='7A9CC0', italic=True, size=9, name='Calibri')
                continue
            ws_ts.cell(row=row, column=2).value = tech['current']
            ws_ts.cell(row=row, column=3).value = tech['rsi']
            ws_ts.cell(row=row, column=4).value = tech['sma20']
            ws_ts.cell(row=row, column=5).value = tech['sma50'] or ''
            ws_ts.cell(row=row, column=6).value = tech['pct_from_sma20']
            ws_ts.cell(row=row, column=6).number_format = '0.00%' if False else '0.00'
            ws_ts.cell(row=row, column=7).value = tech['lower_band']
            ws_ts.cell(row=row, column=8).value = tech['bb_position']
            ws_ts.cell(row=row, column=9).value = tech['volume_ratio']
            ws_ts.cell(row=row, column=10).value = tech['change_5d']
            # Signal and color coding
            rsi = tech['rsi']
            bb = tech['bb_position']
            if rsi < 30 or bb < 0.10:
                signal = 'BUY PUT SIGNAL'
                sig_color = '00E5CC'
                row_color = '0D2A1A'
            elif rsi < 35 or bb < 0.20:
                signal = 'WATCH'
                sig_color = 'FFB347'
                row_color = '1A1A0A'
            elif rsi > 70 or bb > 0.85:
                signal = 'OVERBOUGHT — AVOID'
                sig_color = 'FF4444'
                row_color = '2A0A0A'
            else:
                signal = 'NEUTRAL'
                sig_color = '7A9CC0'
                row_color = '152236'
            ws_ts.cell(row=row, column=11).value = signal
            ws_ts.cell(row=row, column=11).font = Font(color=sig_color, bold=True, size=10, name='Calibri')
            # Color entire row based on signal
            for col in range(1, 11):
                ws_ts.cell(row=row, column=col).fill = bg(row_color)
            ws_ts.cell(row=row, column=11).fill = bg(row_color)

        print('  TechnicalSnapshot updated (' + str(len(ticker_rows)) + ' tickers)')

        # ── 3. Update MacroCalendar upcoming events ───────────────────
        ws_mc = wb['MacroCalendar']
        today = date.today()
        # Clear existing upcoming rows 5-38
        for row in range(5, 39):
            for col in range(1, 8):
                ws_mc.cell(row=row, column=col).value = None
        # Write fresh events
        high_events = [e for e in macro_events if e['impact'] == 'HIGH']
        med_events = [e for e in macro_events if e['impact'] == 'MEDIUM']
        all_events = sorted(high_events + med_events, key=lambda x: x['date'])
        for i, e in enumerate(all_events[:30]):
            row = 5 + i
            days_away = (e['date'] - today).days
            c_date = ws_mc.cell(row=row, column=1)
            c_date.value = e['date']
            c_date.number_format = 'MMM DD YYYY'
            ws_mc.cell(row=row, column=2).value = e['name']
            c_imp = ws_mc.cell(row=row, column=3)
            c_imp.value = e['impact']
            if e['impact'] == 'HIGH':
                c_imp.font = Font(color='FF4444', bold=True, size=10, name='Calibri')
            else:
                c_imp.font = Font(color='FFB347', bold=True, size=10, name='Calibri')
            ws_mc.cell(row=row, column=4).value = str(days_away) + ' days away'
            ws_mc.cell(row=row, column=7).value = e['note']
            for col in range(1, 8):
                c = ws_mc.cell(row=row, column=col)
                c.fill = bg('152236')
                if not c.font or c.font.color.rgb in ('00000000','FF000000'):
                    c.font = Font(color='FFFFFF', size=10, name='Calibri')
                c.alignment = Alignment(horizontal='left', vertical='center')
            ws_mc.row_dimensions[row].height = 20
        print('  MacroCalendar updated')

        safe_save(wb)
        print('  Excel file saved successfully')
    except Exception as ex:
        print('Warning: Could not update Excel: ' + str(ex))

# ── CORE FUNCTIONS ────────────────────────────────────────────────────
def load_config():
    with open('C:\\TradingBot\\config.json') as f:
        return json.load(f)

def is_market_open_today():
    nyse = mcal.get_calendar('NYSE')
    today = datetime.now().strftime('%Y-%m-%d')
    schedule = nyse.schedule(start_date=today, end_date=today)
    return not schedule.empty

def get_next_friday():
    today = datetime.now()
    days_until_friday = (4 - today.weekday()) % 7
    if days_until_friday == 0:
        days_until_friday = 7
    return (today + pd.Timedelta(days=days_until_friday)).strftime('%Y-%m-%d')

def get_monthly_expiries():
    today = date.today()
    expiries = []
    for months_ahead in [1, 2]:
        year = today.year + (today.month + months_ahead - 1) // 12
        month = (today.month + months_ahead - 1) % 12 + 1
        first = date(year, month, 1)
        day = first
        fridays = []
        while day.month == month:
            if day.weekday() == 4:
                fridays.append(day)
            day += timedelta(days=1)
        if len(fridays) >= 3:
            third_friday = fridays[2]
            dte = (third_friday - today).days
            expiries.append((third_friday.strftime('%Y-%m-%d'), dte))
    return expiries

def clean_header(h):
    if h is None: return None
    return str(h).replace('\n','').replace(' ','')

def get_stock_data(ticker):
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        hist = stock.history(period='1y')
        if hist.empty: return None
        current = info.get('currentPrice') or info.get('regularMarketPrice') or float(hist['Close'].iloc[-1])
        pre_market = info.get('preMarketPrice') or current
        week_high = float(hist['Close'].max())
        prev_close = float(hist['Close'].iloc[-2]) if len(hist) > 1 else current
        return {
            'ticker': ticker,
            'current': round(float(current), 2),
            'pre_market': round(float(pre_market), 2),
            'prev_close': round(float(prev_close), 2),
            'week_high': round(float(week_high), 2),
            'pre_market_change_pct': round((float(pre_market) - float(prev_close)) / float(prev_close) * 100, 2),
            'high_proximity_pct': round((float(week_high) - float(current)) / float(week_high) * 100, 2)
        }
    except Exception as ex:
        print('  Could not fetch ' + ticker + ': ' + str(ex))
        return None

def get_earnings_tickers(watchlist):
    earnings_soon = []
    for ticker in watchlist:
        try:
            cal = yf.Ticker(ticker).calendar
            if cal is not None and not cal.empty:
                earn_date = pd.Timestamp(cal.iloc[0, 0])
                days_away = (earn_date - pd.Timestamp.now()).days
                if 0 <= days_away <= 5:
                    earnings_soon.append((ticker, days_away))
        except:
            pass
    return earnings_soon

def check_sympathy_drop(ticker, peer_group, all_data, threshold):
    peers_down = 0
    peers_checked = 0
    for peer in peer_group:
        if peer == ticker or peer not in all_data or all_data[peer] is None:
            continue
        peers_checked += 1
        if all_data[peer]['pre_market_change_pct'] <= -(threshold * 100):
            peers_down += 1
    return peers_checked >= 2 and peers_down >= 2

def find_put_candidates(cfg, all_data, earnings_tickers, open_positions):
    rules = cfg['rules']
    exclusions = cfg['exclusions']
    tier1 = cfg['tiers']['tier1']
    tier2 = cfg['tiers']['tier2']
    peer_groups = cfg['peer_groups']
    candidates = []
    earnings_list = [e[0] for e in earnings_tickers]
    for ticker, data in all_data.items():
        if data is None or ticker in exclusions or ticker in earnings_list or ticker in open_positions:
            continue
        if ticker not in tier1 and ticker not in tier2:
            continue
        if data['high_proximity_pct'] > rules['high_proximity_pct'] * 100:
            continue
        if data['pre_market_change_pct'] >= -(rules['sympathy_drop_pct'] * 100):
            continue
        ticker_group = None
        for gn, members in peer_groups.items():
            if ticker in members:
                ticker_group = gn
                break
        if ticker_group is None:
            continue
        if not check_sympathy_drop(ticker, peer_groups[ticker_group], all_data, rules['sympathy_drop_pct']):
            continue
        tier = 'Tier 1' if ticker in tier1 else 'Tier 2'
        notional_min = rules['tier1_notional_min'] if ticker in tier1 else rules['tier2_notional_min']
        notional_max = rules['tier1_notional_max'] if ticker in tier1 else rules['tier2_notional_max']
        avg_notional = (notional_min + notional_max) / 2
        strike_low = round(data['current'] * (1 - rules['put_strike_max_otm']), 2)
        strike_high = round(data['current'] * (1 - rules['put_strike_min_otm']), 2)
        contracts = max(1, round(avg_notional / (data['current'] * 100)))
        est_prem = round((strike_low + strike_high) / 2 * rules['min_weekly_premium_pct'] * contracts * 100, 2)
        candidates.append({
            'ticker': ticker, 'tier': tier, 'group': ticker_group,
            'current': data['current'], 'pre_market': data['pre_market'],
            'drop_pct': data['pre_market_change_pct'],
            'week_high': data['week_high'], 'proximity_pct': data['high_proximity_pct'],
            'strike_low': strike_low, 'strike_high': strike_high,
            'contracts': contracts, 'est_premium': est_prem,
            'est_per_contract': round(est_prem / contracts, 2)
        })
    return candidates

def find_longer_dated_candidates(cfg, all_data, earnings_tickers, open_positions):
    rules = cfg['rules']
    exclusions = cfg['exclusions']
    peer_groups = cfg['peer_groups']
    longer_tier1 = ['MSFT', 'AMZN', 'GOOGL', 'META', 'NVDA', 'AVGO']
    candidates = []
    earnings_list = [e[0] for e in earnings_tickers]
    expiries = get_monthly_expiries()
    for ticker, data in all_data.items():
        if data is None or ticker in exclusions or ticker in earnings_list or ticker in open_positions:
            continue
        if ticker not in longer_tier1:
            continue
        if data['high_proximity_pct'] > rules['high_proximity_pct'] * 100:
            continue
        if data['pre_market_change_pct'] >= -(rules['sympathy_drop_pct'] * 100):
            continue
        ticker_group = None
        for gn, members in peer_groups.items():
            if ticker in members:
                ticker_group = gn
                break
        if ticker_group is None:
            continue
        if not check_sympathy_drop(ticker, peer_groups[ticker_group], all_data, rules['sympathy_drop_pct']):
            continue
        current = data['current']
        exp_results = []
        for exp_date, dte in expiries:
            if dte < 25 or dte > 65:
                continue
            weeks = round(dte / 7, 1)
            delta_proxy = 0.12 + (dte / 71) * 0.08
            strike = round(current * (1 - delta_proxy), 2)
            notional = round(strike * 100, 2)
            time_factor = (dte / 365) ** 0.5
            est_prem = round(current * 0.20 * time_factor * 0.20, 2)
            roi = round(est_prem / strike * 100, 2)
            exp_results.append({'exp_date':exp_date,'dte':dte,'weeks':weeks,'strike':strike,
                                 'notional':notional,'est_prem':est_prem,'roi':roi,
                                 'weekly_roi':round(roi/weeks,2)})
        if not exp_results:
            continue
        candidates.append({'ticker':ticker,'group':ticker_group,'current':current,
                           'drop_pct':data['pre_market_change_pct'],'week_high':data['week_high'],
                           'proximity_pct':data['high_proximity_pct'],'expiries':exp_results})
    return candidates

def find_mean_reversion_candidates(cfg, all_data, earnings_tickers, open_positions, tech_cache):
    tier1 = cfg['tiers']['tier1']
    tier2 = cfg['tiers']['tier2']
    exclusions = cfg['exclusions']
    earnings_list = [e[0] for e in earnings_tickers]
    candidates = []
    for ticker in tier1 + tier2:
        if ticker in exclusions or ticker in earnings_list or ticker in open_positions:
            continue
        if all_data.get(ticker) is None:
            continue
        data = all_data[ticker]
        if data['high_proximity_pct'] > cfg['rules']['high_proximity_pct'] * 100:
            continue
        # Check not a sympathy drop
        ticker_group = None
        for gn, members in cfg['peer_groups'].items():
            if ticker in members:
                ticker_group = gn
                break
        if ticker_group:
            peers_down = sum(1 for peer in cfg['peer_groups'][ticker_group]
                            if peer != ticker and all_data.get(peer) and
                            all_data[peer]['pre_market_change_pct'] < -2.0)
            if peers_down >= 2:
                continue
        tech = tech_cache.get(ticker)
        if tech is None:
            continue
        score = 0
        if tech['rsi'] < 25: score += 3
        elif tech['rsi'] < 30: score += 2
        elif tech['rsi'] < 35: score += 1
        elif tech['rsi'] > 70: continue
        if tech['bb_position'] < 0.05: score += 3
        elif tech['bb_position'] < 0.15: score += 2
        elif tech['bb_position'] < 0.25: score += 1
        if tech['pct_from_sma20'] < -10: score += 2
        elif tech['pct_from_sma20'] < -5: score += 1
        if tech['volume_ratio'] < 0.7: score += 1
        elif tech['volume_ratio'] > 2.0: score -= 1
        if tech['change_5d'] < -10: score += 1
        if score < 2:
            continue
        conviction = 'HIGH CONVICTION' if score >= 5 else 'MODERATE' if score >= 3 else 'WATCH'
        candidates.append({
            'ticker': ticker, 'conviction': conviction, 'score': score, 'tech': tech,
            'weekly_strike': round(tech['current'] * 0.95, 2),
            'monthly_strike': round(tech['current'] * 0.88, 2)
        })
    return sorted(candidates, key=lambda x: x['score'], reverse=True)

def load_positions():
    try:
        wb = openpyxl.load_workbook(POSITIONS_FILE, read_only=True, data_only=True)
        puts_ws = wb['OpenPuts']
        assigned_ws = wb['AssignedPositions']
        open_puts = []
        assigned = []
        buy_writes = []
        puts_headers = [clean_header(c.value) for c in puts_ws[3]]
        for row in puts_ws.iter_rows(min_row=4, values_only=True):
            if row[0] is None: continue
            put = {h: row[i] for i, h in enumerate(puts_headers) if h}
            open_puts.append(put)
        assigned_headers = [clean_header(c.value) for c in assigned_ws[3]]
        for row in assigned_ws.iter_rows(min_row=4, values_only=True):
            if row[0] is None: continue
            pos = {h: row[i] for i, h in enumerate(assigned_headers) if h}
            assigned.append(pos)
        if 'BuyWrites' in wb.sheetnames:
            bw_ws = wb['BuyWrites']
            bw_headers = [clean_header(c.value) for c in bw_ws[4]]
            for row in bw_ws.iter_rows(min_row=5, values_only=True):
                if row[0] is None: continue
                
                return open_puts, assigned, buy_writes
    except Exception as ex:
        print('Warning: Could not read positions.xlsx: ' + str(ex))
        return [], [], []
bw = {h: row[i] for i, h in enumerate(bw_headers) if h}
                # Skip summary / label rows — real positions have a positive share count.
                sh = bw.get('Shares')
                try:
                    if sh is None or float(sh) <= 0:
                        continue
                except (TypeError, ValueError):
                    continue
                buy_writes.append(bw)
def check_stops(assigned, all_data):
    alerts = []
    for pos in assigned:
        ticker = pos['Ticker']
        cost_basis = float(pos['CostBasis'])
        highest = float(pos.get('HighestPriceSeen') or cost_basis)
        if ticker not in all_data or all_data[ticker] is None: continue
        current = all_data[ticker]['current']
        static_stop = round(cost_basis * 0.95, 2)
        trailing_active = current >= cost_basis * 1.10 or highest >= cost_basis * 1.10
        trailing_stop = round(highest * 0.95, 2)
        stop_price = trailing_stop if trailing_active else static_stop
        stop_type = 'TRAILING' if trailing_active else 'STATIC'
        pnl_pct = round((current - cost_basis) / cost_basis * 100, 2)
        has_cc = pos.get('CoveredCallStrike') is not None
        status = 'OK'
        if current <= stop_price:
            status = 'STOP HIT - BUY BACK CALL FIRST' if has_cc else 'STOP HIT - SELL SHARES'
        elif current <= stop_price * 1.03:
            status = 'APPROACHING STOP - WATCH CLOSELY'
        alerts.append({'ticker':ticker,'current':current,'cost_basis':cost_basis,'pnl_pct':pnl_pct,
                       'stop_price':stop_price,'stop_type':stop_type,'trailing_active':trailing_active,
                       'highest':highest,'has_covered_call':has_cc,
                       'covered_call_strike':pos.get('CoveredCallStrike'),
                       'covered_call_expiry':pos.get('CoveredCallExpiry'),'status':status})
    return alerts

def get_call_recommendations(assigned, all_data, rules):
    recommendations = []
    for pos in assigned:
        ticker = pos['Ticker']
        cost_basis = float(pos['CostBasis'])
        shares = int(pos['Shares'])
        highest = float(pos.get('HighestPriceSeen') or cost_basis)
        if pos.get('CoveredCallStrike') is not None: continue
        if ticker not in all_data or all_data[ticker] is None: continue
        current = all_data[ticker]['current']
        trailing_active = highest >= cost_basis * 1.10
        stop_price = round(highest * 0.95, 2) if trailing_active else round(cost_basis * 0.95, 2)
        pnl_pct = round((current - cost_basis) / cost_basis * 100, 2)
        if current <= stop_price * 1.03: continue
        if current >= cost_basis:
            mode = 'NORMAL'
            call_strike = round(current * (1 + rules['call_strike_min_otm']), 2)
            call_strike_high = round(current * (1 + rules['call_strike_max_otm']), 2)
            est_premium = round(current * rules['min_call_premium_pct'], 2)
        elif current >= cost_basis * (1 - rules['recovery_mode_threshold']):
            mode = 'RECOVERY'
            call_strike = round(current, 2)
            call_strike_high = round(current * 1.01, 2)
            est_premium = round(current * rules['min_call_premium_pct'] * 1.5, 2)
        else:
            continue
        recommendations.append({'ticker':ticker,'current':current,'cost_basis':cost_basis,
                                 'pnl_pct':pnl_pct,'shares':shares,'mode':mode,
                                 'call_strike':call_strike,'call_strike_high':call_strike_high,
                                 'est_premium':est_premium,'est_total_premium':round(est_premium*shares,2),
                                 'stop_price':stop_price,'stop_type':'TRAILING' if trailing_active else 'STATIC'})
    return recommendations

# ── REPORT BUILDERS ───────────────────────────────────────────────────
def build_market_regime_section(vix, spy_trend, macro_events):
    L = []
    L.append('=' * 60)
    L.append('SECTION 0 - DAILY MARKET REGIME')
    L.append('-' * 40)
    today = date.today()
    next_7 = [e for e in macro_events if e['impact'] == 'HIGH' and (e['date'] - today).days <= 7]
    if vix:
        if vix < 15: vix_text = 'LOW (' + str(vix) + ') — premiums thin'
        elif vix < 20: vix_text = 'NORMAL (' + str(vix) + ') — standard environment'
        elif vix < 30: vix_text = 'ELEVATED (' + str(vix) + ') — GOOD for premium selling'
        else: vix_text = 'EXTREME (' + str(vix) + ') — rich premiums but high risk'
        L.append('VIX Fear Gauge:   ' + vix_text)
    else:
        L.append('VIX:              Could not fetch')
    if spy_trend:
        L.append('SPY Trend:        ' + spy_trend['trend'] + ' | $' + str(spy_trend['current']) + ' | 1d: ' + str(spy_trend['change_1d']) + '% | 5d: ' + str(spy_trend['change_5d']) + '%')
        L.append('                  20d SMA: $' + str(spy_trend['sma20']) + ' | 50d SMA: $' + str(spy_trend['sma50']))
    if next_7:
        L.append('Macro Risk:       ' + str(len(next_7)) + ' HIGH IMPACT event(s) within 7 days')
        for e in next_7:
            days = (e['date'] - today).days
            label = 'TODAY' if days == 0 else 'TOMORROW' if days == 1 else str(days) + ' days'
            L.append('                  ' + e['name'] + ' (' + label + ')')
    else:
        L.append('Macro Risk:       No high impact events in next 7 days')
    verdict, reasons = get_market_regime(vix, spy_trend)
    L.append('')
    L.append('OVERALL VERDICT:  *** ' + verdict + ' FOR PUT WRITING ***')
    for r in reasons:
        L.append('  - ' + r)
    L.append('')
    return '\n'.join(L)

def format_candidate_analysis(ticker, current, strike_low, contracts, analyst, iv_data, earnings_dt, earnings_days, macro_warn, tech, all_data, cfg):
    L = []
    verdict, reasons_for, reasons_against, score = generate_recommendation(
        ticker, tech, analyst, iv_data, earnings_dt, earnings_days, macro_warn, cfg)
    actual_notional = contracts * current * 100
    if analyst:
        if analyst.get('target'):
            L.append('Analyst target:   $' + str(analyst['target']) + ' (' + str(analyst['upside']) + '% upside) — ' + str(analyst['rec']) + ' (' + str(analyst['n_analysts']) + ' analysts)')
        if analyst.get('beta_text'):
            L.append('Beta:             ' + analyst['beta_text'])
        if analyst.get('fwd_pe'):
            L.append('Forward P/E:      ' + str(analyst['fwd_pe']))
    if iv_data and iv_data.get('iv_rank'):
        L.append('IV vs History:    ' + iv_data['iv_rank'])
    if earnings_days is not None:
        L.append('Earnings:         ' + str(earnings_dt) + ' (' + str(earnings_days) + ' days away)')
    if tech:
        breakeven = round(strike_low - (contracts * current * 0.01) / 100, 2)
        L.append('Breakeven:        $' + str(breakeven) + ' (approx if assigned at lower strike)')
    L.append('')
    L.append('*** RECOMMENDATION: ' + verdict + ' ***')
    for r in reasons_for:
        L.append('  + ' + r)
    for r in reasons_against:
        L.append('  - ' + r)
    return L, verdict

def build_report(cfg, candidates, earnings_tickers, all_data, watchlist_flags, open_puts, assigned, buy_writes, macro_events, vix):
    now = datetime.now().strftime('%A %B %d, %Y %I:%M %p')
    L = []
    L.append('TRADING BOT MORNING REPORT')
    L.append(now)
    L.append('')
    L.append('SECTION A - PORTFOLIO SNAPSHOT')
    L.append('-' * 40)
    total = cfg['portfolio']['total_value']
    reserve = cfg['portfolio']['dry_powder_reserve']
    committed = sum(float(p.get('Strike',0)) * int(p.get('Contracts',1)) * 100 for p in open_puts if p.get('Strike'))
    assigned_value = sum(float(a.get('CostBasis',0)) * int(a.get('Shares',0)) for a in assigned if a.get('CostBasis'))
    # BuyWrites capital = PurchasePrice x Shares (use raw columns, not formula columns)
    bw_value = 0
    bw_premium = 0
    for bw in buy_writes:
        try:
            price = bw.get('PurchasePrice') or bw.get('Purchase\nPrice') or 0
            shares = bw.get('Shares') or 0
            # Skip if either is a formula string
            if isinstance(price, str) or isinstance(shares, str):
                continue
            bw_value += float(price) * int(shares)
            prem = bw.get('CallPremium') or bw.get('Call\nPremium') or 0
            if not isinstance(prem, str):
                bw_premium += float(prem)
        except:
            pass
    # Net capital actually at risk in a buy-write = stock cost - premium collected.
    # Covered-call premium is cash received, so it offsets the buying power consumed.
    bw_net = bw_value - bw_premium
    total_deployed = committed + assigned_value + bw_net
    available = total - reserve - total_deployed
    L.append('Total portfolio value:    $' + format(int(total), ','))
    L.append('Dry powder reserve:       $' + format(int(reserve), ','))
    L.append('Cash in open puts:        $' + format(int(committed), ','))
    L.append('Cash in buy-write stocks: $' + format(int(bw_value), ','))
    if bw_premium:
        L.append('  less call premium recd: -$' + format(int(bw_premium), ','))
    L.append('Cash in assigned stocks:  $' + format(int(assigned_value), ','))
    L.append('Total deployed:           $' + format(int(total_deployed), ','))
    surplus = total - total_deployed  # cash above committed capital, before reserve
    if total_deployed > total:
        L.append('Available for new trades: $0')
        L.append('*** WARNING: OVER-DEPLOYED by $' + format(int(total_deployed - total), ',') +
                 ' — deployed capital exceeds account value. Update total_value in config.json, or verify positions. ***')
    else:
        free_for_new = surplus - reserve
        L.append('Available for new trades: $' + format(int(free_for_new), ','))
        if free_for_new < 0:
            L.append('NOTE: Fully deployed. $' + format(int(surplus), ',') + ' sits above committed capital, below your $' +
                     format(int(reserve), ',') + ' dry-powder target. Not a rule violation — just no spare cash for new ' +
                     'trades until expiries free up capital.')
        elif free_for_new < reserve:
            L.append('NOTE: Limited dry powder — $' + format(int(free_for_new), ',') + ' available beyond your reserve.')
    # Show buy-write expiry summary
    if buy_writes:
        expiring_soon = []
        from datetime import date as _date
        today = _date.today()
        for bw in buy_writes:
            exp = bw.get('CallExpiry')
            if exp:
                exp_date = exp.date() if hasattr(exp, 'date') else exp
                days_left = (exp_date - today).days
                if days_left <= 7:
                    capital = float(bw.get('PurchasePrice',0)) * int(bw.get('Shares',0)) if not isinstance(bw.get('PurchasePrice'), str) else 0
                    expiring_soon.append((bw.get('Ticker','?'), days_left, exp_date, capital))
        if expiring_soon:
            total_freeing = sum(c for _,_,_,c in expiring_soon)
            L.append('')
            L.append('BUY-WRITE POSITIONS EXPIRING WITHIN 7 DAYS:')
            for ticker, days, exp_dt, cap in expiring_soon:
                label = 'FRIDAY' if days <= 4 else str(days) + ' days'
                L.append('  ' + ticker + ' expires ' + label + ' (' + str(exp_dt) + ') — frees $' + format(int(cap), ','))
            L.append('  Total capital freeing up: $' + format(int(total_freeing), ','))
    L.append('Note: Update total_value in config.json as it changes.')
    L.append('')
    # Buy-write open positions summary
    if buy_writes:
        L.append('BUY-WRITE OPEN POSITIONS')
        L.append('-' * 40)
        total_bw_gain = 0
        for bw in buy_writes:
            try:
                ticker = str(bw.get('Ticker',''))
                shares = int(bw.get('Shares') or 0)
                strike = float(bw.get('CallStrike') or bw.get('Call\nStrike') or 0)
                expiry = bw.get('CallExpiry') or bw.get('Call\nExpiry') or 'N/A'
                expiry_str = expiry.strftime('%Y-%m-%d') if hasattr(expiry, 'strftime') else str(expiry)
                prem = float(bw.get('CallPremium') or bw.get('Call\nPremium') or 0)
                purchase = float(bw.get('PurchasePrice') or bw.get('Purchase\nPrice') or 0)
                net_basis = purchase - (prem / shares) if shares > 0 else purchase
                max_gain = (strike - net_basis) * shares if strike > 0 else 0
                total_bw_gain += max_gain
                current_data = all_data.get(ticker)
                current_str = '$' + str(current_data['current']) if current_data else 'N/A'
                buffer = round((current_data['current'] - strike) / strike * 100, 1) if current_data and current_data.get('current') else None
                buffer_str = (str(buffer) + '% above call strike') if buffer is not None else 'N/A'
                L.append(ticker + ' | ' + str(shares) + ' shares @ $' + str(round(purchase,2)) + ' | Call: $' + str(strike) + ' exp ' + expiry_str)
                L.append('  Current: ' + current_str + ' | ' + buffer_str + ' | Max gain: $' + str(round(max_gain, 2)))
                L.append('')
            except Exception as ex:
                L.append(str(bw.get('Ticker','?')) + ' | Error reading: ' + str(ex))
        L.append('Total max gain if all called away: $' + format(round(total_bw_gain, 2), ','))
        L.append('')
    L.append('SECTION B - WEEKLY PUT CANDIDATES TODAY')
    L.append('-' * 40)
    today = date.today()
    next_7_high = [e for e in macro_events if e['impact'] == 'HIGH' and (e['date'] - today).days <= 7]
    macro_warning = next_7_high[0]['name'] if next_7_high else None
    if not candidates:
        L.append('No qualifying weekly put candidates found today.')
        L.append('Reasons: no sympathy drops, earnings soon, or too far from 52W high.')
    else:
        for c in candidates:
            print('  Full analysis for B candidate ' + c['ticker'] + '...')
            analyst = get_analyst_data(c['ticker'])
            iv_data = get_iv_data(c['ticker'])
            earnings_dt, earnings_days = get_earnings_date(c['ticker'])
            tech = calculate_technicals(c['ticker'])
            actual_notional = c['contracts'] * c['current'] * 100
            L.append('Ticker:           ' + c['ticker'] + ' (' + c['tier'] + ' | ' + c['group'] + ')')
            L.append('Current price:    $' + str(c['current']) + ' | Drop today: ' + str(c['drop_pct']) + '%')
            L.append('52W high:         $' + str(c['week_high']) + ' (' + str(c['proximity_pct']) + '% below)')
            L.append('Strike range:     $' + str(c['strike_low']) + ' to $' + str(c['strike_high']))
            L.append('Contracts:        ' + str(c['contracts']) + ' | Notional: $' + format(int(actual_notional), ','))
            L.append('Est. premium:     $' + str(c['est_premium']) + ' total')
            analysis_lines, verdict = format_candidate_analysis(
                c['ticker'], c['current'], c['strike_low'], c['contracts'],
                analyst, iv_data, earnings_dt, earnings_days, macro_warning, tech, all_data, cfg)
            L.extend(analysis_lines)
            L.append('')
            L.append('*** Verify premium and delta in ATP before placing order ***')
            L.append('')
        L.append('ATP ORDER TICKETS')
        L.append('-' * 40)
        next_friday = get_next_friday()
        for c in candidates:
            L.append('SELL TO OPEN PUT | ' + c['ticker'] + ' | Expiry: ' + next_friday + ' | Strike: $' + str(c['strike_low']) + '-$' + str(c['strike_high']) + ' | Contracts: ' + str(c['contracts']))
        L.append('')
    L.append('SECTION E - WATCHLIST FLAGS')
    L.append('-' * 40)
    if earnings_tickers:
        L.append('Earnings within 5 days - DO NOT write puts on these:')
        for ticker, days in earnings_tickers:
            L.append('  ' + ticker + ': ' + str(days) + ' day(s) away')
    else:
        L.append('No earnings alerts today.')
    L.append('')
    if watchlist_flags:
        L.append('Large pre-market moves (>5%):')
        for flag in watchlist_flags:
            direction = 'UP' if flag['change'] > 0 else 'DOWN'
            L.append('  ' + flag['ticker'] + ': ' + str(flag['change']) + '% ' + direction)
    else:
        L.append('No large pre-market moves today.')
    L.append('')
    return '\n'.join(L)

def build_longer_dated_section(candidates, macro_events):
    L = []
    L.append('SECTION B2 - LONGER DATED PUT CANDIDATES (30-60 DTE)')
    L.append('-' * 40)
    L.append('Tier 1 names only: MSFT AMZN GOOGL META NVDA AVGO | Delta ~0.20 | 12-20% OTM')
    L.append('')
    if not candidates:
        L.append('No qualifying longer-dated candidates today.')
        return '\n'.join(L)
    today = date.today()
    for c in candidates:
        print('  Full analysis for B2 candidate ' + c['ticker'] + '...')
        analyst = get_analyst_data(c['ticker'])
        iv_data = get_iv_data(c['ticker'])
        earnings_dt, earnings_days = get_earnings_date(c['ticker'])
        tech = calculate_technicals(c['ticker'])
        next_high = [e for e in macro_events if e['impact'] == 'HIGH' and (e['date'] - today).days <= 7]
        macro_warn = next_high[0]['name'] if next_high else None
        L.append('Ticker: ' + c['ticker'] + ' | Price: $' + str(c['current']) + ' | Drop: ' + str(c['drop_pct']) + '% | 52W high: $' + str(c['week_high']))
        if analyst and analyst.get('target'):
            L.append('Analyst: $' + str(analyst['target']) + ' target | ' + str(analyst['upside']) + '% upside | ' + str(analyst['rec']))
        if analyst and analyst.get('beta_text'):
            L.append('Beta: ' + analyst['beta_text'])
        if iv_data and iv_data.get('iv_rank'):
            L.append('IV: ' + iv_data['iv_rank'])
        if earnings_days is not None:
            L.append('Earnings: ' + str(earnings_dt) + ' (' + str(earnings_days) + ' days)')
        verdict, reasons_for, reasons_against, score = generate_recommendation(
            c['ticker'], tech, analyst, iv_data, earnings_dt, earnings_days, macro_warn, {})
        L.append('*** RECOMMENDATION: ' + verdict + ' ***')
        for r in reasons_for: L.append('  + ' + r)
        for r in reasons_against: L.append('  - ' + r)
        L.append('')
        for e in c['expiries']:
            L.append('  ' + e['exp_date'] + ' (' + str(e['dte']) + ' DTE) | Strike: $' + str(e['strike']) + ' | Est prem: $' + str(e['est_prem']) + ' | Weekly ROI: ' + str(e['weekly_roi']) + '%')
        if c['expiries']:
            best = c['expiries'][0]
            L.append('  ORDER: SELL PUT | ' + c['ticker'] + ' | ' + best['exp_date'] + ' | $' + str(best['strike']) + ' | Verify delta ~0.20 in ATP')
        L.append('')
        L.append('-' * 40)
    return '\n'.join(L)

def build_mean_reversion_section(candidates, macro_events):
    L = []
    L.append('SECTION G - MEAN REVERSION CANDIDATES')
    L.append('-' * 40)
    L.append('Isolated drops — no peer confirmation. Technical bounce candidates.')
    L.append('')
    if not candidates:
        L.append('No mean reversion candidates today.')
        return '\n'.join(L)
    today = date.today()
    next_friday = get_next_friday()
    for c in candidates:
        tech = c['tech']
        print('  Full analysis for G candidate ' + c['ticker'] + '...')
        analyst = get_analyst_data(c['ticker'])
        iv_data = get_iv_data(c['ticker'])
        earnings_dt, earnings_days = get_earnings_date(c['ticker'])
        next_high = [e for e in macro_events if e['impact'] == 'HIGH' and (e['date'] - today).days <= 7]
        macro_warn = next_high[0]['name'] if next_high else None
        verdict, reasons_for, reasons_against, score = generate_recommendation(
            c['ticker'], tech, analyst, iv_data, earnings_dt, earnings_days, macro_warn, {})
        L.append('*** ' + c['ticker'] + ' — ' + c['conviction'] + ' (Technical Score: ' + str(c['score']) + '/8) ***')
        L.append('Price: $' + str(tech['current']) + ' | RSI: ' + str(tech['rsi']) + (' OVERSOLD' if tech['rsi'] < 30 else '') + ' | BB: ' + str(tech['bb_position']) + ' | vs SMA20: ' + str(tech['pct_from_sma20']) + '%')
        L.append('ATR: $' + str(tech['atr']) + ' (' + str(tech['atr_pct']) + '%) | Vol: ' + str(tech['volume_ratio']) + 'x avg | 1d: ' + str(tech['change_1d']) + '% | 5d: ' + str(tech['change_5d']) + '%')
        if analyst:
            if analyst.get('beta_text'):
                L.append('Beta: ' + analyst['beta_text'])
            if analyst.get('target'):
                L.append('Analyst: $' + str(analyst['target']) + ' target (' + str(analyst['upside']) + '% upside) | ' + str(analyst['rec']) + ' | ' + str(analyst['n_analysts']) + ' analysts')
                assign_price = c['weekly_strike']
                if analyst['target'] and analyst['target'] > assign_price:
                    L.append('If assigned at $' + str(assign_price) + ': analysts target ' + str(round((analyst['target'] - assign_price) / assign_price * 100, 1)) + '% recovery')
                elif analyst['target'] and analyst['target'] < assign_price:
                    L.append('WARNING: Analyst target $' + str(analyst['target']) + ' is BELOW assignment price $' + str(assign_price))
        if iv_data and iv_data.get('iv_rank'):
            L.append('IV: ' + iv_data['iv_rank'])
        if earnings_days is not None:
            L.append('Earnings: ' + str(earnings_dt) + ' (' + str(earnings_days) + ' days)')
        L.append('')
        L.append('*** RECOMMENDATION: ' + verdict + ' ***')
        for r in reasons_for: L.append('  + ' + r)
        for r in reasons_against: L.append('  - ' + r)
        L.append('')
        L.append('Strikes: Weekly (' + next_friday + '): $' + str(c['weekly_strike']) + ' | Monthly: $' + str(c['monthly_strike']))
        L.append('*** Verify delta, premium and IV in ATP before trading ***')
        L.append('')
        L.append('-' * 40)
    return '\n'.join(L)

def build_sections_cd(open_puts, assigned, all_data, rules):
    L = []
    L.append('SECTION C - COVERED CALL OPPORTUNITIES')
    L.append('-' * 40)
    if not assigned:
        L.append('No assigned positions on file.')
    else:
        call_recs = get_call_recommendations(assigned, all_data, rules)
        if not call_recs:
            L.append('No covered call opportunities today.')
        else:
            next_friday = get_next_friday()
            for r in call_recs:
                L.append(r['ticker'] + ' | Price: $' + str(r['current']) + ' | Basis: $' + str(r['cost_basis']) + ' | PnL: ' + str(r['pnl_pct']) + '% | Mode: ' + r['mode'])
                L.append('  Strike: $' + str(r['call_strike']) + '-$' + str(r['call_strike_high']) + ' | Est prem: $' + str(r['est_premium']) + '/share | Total: $' + str(r['est_total_premium']))
                L.append('  Stop: $' + str(r['stop_price']) + ' (' + r['stop_type'] + ')')
                L.append('  ORDER: SELL CALL | ' + r['ticker'] + ' | Expiry: ' + next_friday + ' | Contracts: ' + str(int(r['shares']/100)))
                L.append('')
    L.append('SECTION D - STOP ALERTS AND POSITION STATUS')
    L.append('-' * 40)
    if not assigned:
        L.append('No assigned positions on file.')
    else:
        stop_alerts = check_stops(assigned, all_data)
        has_alerts = any(a['status'] != 'OK' for a in stop_alerts)
        for a in stop_alerts:
            if a['status'] != 'OK':
                L.append('*** ' + a['status'] + ' ***')
                L.append(a['ticker'] + ' | Current: $' + str(a['current']) + ' | Stop: $' + str(a['stop_price']) + ' | PnL: ' + str(a['pnl_pct']) + '%')
                if a['has_covered_call']:
                    L.append('WARNING: BUY BACK CALL BEFORE SELLING SHARES')
                L.append('')
        if not has_alerts:
            L.append('No stop alerts. All positions within normal range.')
            L.append('')
            for a in stop_alerts:
                L.append(a['ticker'] + ' | Current: $' + str(a['current']) + ' | Stop: $' + str(a['stop_price']) + ' | PnL: ' + str(a['pnl_pct']) + '%')
    L.append('')
    L.append('OPEN PUTS SUMMARY')
    L.append('-' * 40)
    if not open_puts:
        L.append('No open put positions on file.')
    else:
        for p in open_puts:
            prem = p.get('PremiumCollected', 'N/A')
            try: prem_str = '$' + str(round(float(prem), 2))
            except: prem_str = 'N/A'
            expiry = p.get('Expiry', 'N/A')
            expiry_str = expiry.strftime('%Y-%m-%d') if hasattr(expiry, 'strftime') else str(expiry)
            ticker = str(p['Ticker'])
            strike = float(p['Strike'])
            try: breakeven = round(strike - float(prem) / 100, 2)
            except: breakeven = 'N/A'
            current_data = all_data.get(ticker)
            current_str = '$' + str(current_data['current']) if current_data else 'N/A'
            if current_data and current_data.get('current'):
                buffer = round((current_data['current'] - strike) / strike * 100, 1)
                buffer_str = str(buffer) + '% above strike'
            else:
                buffer_str = 'N/A'
            macro_flag = p.get('MacroRiskFlag', '')
            L.append(ticker + ' | Strike: $' + str(strike) + ' | Expiry: ' + expiry_str + ' | Premium: ' + prem_str)
            L.append('  Current: ' + current_str + ' | Buffer: ' + buffer_str + ' | Breakeven: $' + str(breakeven))
            if macro_flag and macro_flag != 'None' and macro_flag != '':
                L.append('  Macro events before expiry: ' + str(macro_flag))
    return '\n'.join(L)

def find_buy_write_candidates(cfg, all_data, earnings_tickers, tech_cache):
    tier1 = cfg['tiers']['tier1']
    tier2 = cfg['tiers']['tier2']
    exclusions = cfg['exclusions']
    earnings_list = [e[0] for e in earnings_tickers]
    candidates = []
    for ticker in tier1 + tier2:
        if ticker in exclusions or ticker in earnings_list:
            continue
        data = all_data.get(ticker)
        if data is None:
            continue
        # Must be within 15% of 52W high — quality filter
        if data['high_proximity_pct'] > 15:
            continue
        current = data['current']
        try:
            stock = yf.Ticker(ticker)
            opt_dates = stock.options
            if not opt_dates:
                continue
            # Use nearest weekly expiry
            chain = stock.option_chain(opt_dates[0])
            calls = chain.calls
            if calls.empty:
                continue
            # Scan ALL ITM calls (strikes below current price, from 0.5% to 8% below)
            # This mirrors the real approach — find the strike with the best net return
            best_candidate = None
            best_net_pct = 0.0
            for _, call_row in calls.iterrows():
                strike = float(call_row['strike'])
                # Only consider ITM calls (below current price)
                if strike >= current:
                    continue
                # Only consider strikes between 0.5% and 8% below current
                discount_pct = (current - strike) / current
                if discount_pct < 0.005 or discount_pct > 0.08:
                    continue
                bid = float(call_row.get('bid', 0) or 0)
                ask = float(call_row.get('ask', 0) or 0)
                iv = float(call_row.get('impliedVolatility', 0) or 0)
                if bid <= 0 or ask <= 0:
                    continue
                mid = round((bid + ask) / 2, 2)
                # Skip if spread is too wide (> 20% of mid) — liquidity concern
                if (ask - bid) / mid > 0.20:
                    continue
                intrinsic = max(0, current - strike)
                time_value = max(0, mid - intrinsic)
                stock_loss = strike - current  # negative number
                net_per_share = mid + stock_loss
                net_pct = net_per_share / current
                # Must meet 1% weekly minimum
                if net_pct < 0.01:
                    continue
                if net_pct > best_net_pct:
                    best_net_pct = net_pct
                    best_candidate = {
                        'ticker': ticker,
                        'current': current,
                        'call_strike': strike,
                        'strike_discount_pct': round(discount_pct * 100, 2),
                        'call_premium_mid': mid,
                        'intrinsic': round(intrinsic, 2),
                        'time_value': round(time_value, 2),
                        'stock_loss_per_share': round(stock_loss, 2),
                        'net_per_share': round(net_per_share, 2),
                        'net_pct': round(net_pct * 100, 2),
                        'expiry': opt_dates[0],
                        'iv': round(iv * 100, 1),
                        'proximity_pct': data['high_proximity_pct'],
                        'week_high': data['week_high'],
                        'bid': bid,
                        'ask': ask,
                    }
            if best_candidate is None:
                continue
            # Enrich with analyst and technical data
            best_candidate['analyst'] = get_analyst_data(ticker)
            best_candidate['tech'] = tech_cache.get(ticker)
            best_candidate['earnings_dt'], best_candidate['earnings_days'] = get_earnings_date(ticker)
            candidates.append(best_candidate)
        except Exception as ex:
            pass
    # ── Risk-adjusted scoring ─────────────────────────────────────
    # Score each candidate on return AND risk — higher is better
    for c in candidates:
        score = 0
        analyst = c.get('analyst', {}) or {}
        tech = c.get('tech') or {}

        # Return component (0-40 points)
        net_pct = c['net_pct']
        if net_pct >= 3.0:   score += 40
        elif net_pct >= 2.0: score += 30
        elif net_pct >= 1.5: score += 20
        elif net_pct >= 1.0: score += 10

        # Time value component — more time value = more cushion (0-10 points)
        tv_pct = c['time_value'] / c['current'] * 100 if c['current'] > 0 else 0
        if tv_pct >= 1.0:   score += 10
        elif tv_pct >= 0.5: score += 5

        # Beta risk penalty (0 to -20 points)
        beta_risk = analyst.get('beta_risk', 'MODERATE')
        if beta_risk == 'VERY HIGH':   score -= 20
        elif beta_risk == 'HIGH':      score -= 10
        elif beta_risk == 'LOW':       score += 5

        # Technical risk (RSI) — avoid overbought (0 to -15 points)
        rsi = tech.get('rsi', 50) if tech else 50
        if rsi > 75:        score -= 15
        elif rsi > 65:      score -= 8
        elif rsi < 35:      score += 8  # oversold = good support
        elif rsi < 45:      score += 4

        # BB Position — avoid near upper band (0 to -10 points)
        bb = tech.get('bb_position', 0.5) if tech else 0.5
        if bb > 0.85:       score -= 10
        elif bb > 0.70:     score -= 5
        elif bb < 0.30:     score += 5  # near lower band = support

        # Analyst consensus bonus (0 to +10 points)
        rec = analyst.get('rec', '')
        if rec == 'STRONG BUY':  score += 10
        elif rec == 'BUY':       score += 5

        # Earnings penalty
        earnings_days = c.get('earnings_days')
        if earnings_days is not None:
            if earnings_days <= 5:   score -= 50  # automatic disqualifier
            elif earnings_days <= 7: score -= 20
            elif earnings_days <= 14: score -= 10

        # Proximity to 52W high — avoid stocks already very extended
        prox = c.get('proximity_pct', 10)
        if prox < 3:   score -= 10  # within 3% of high, stretched
        elif prox > 10: score += 5  # more room above = safer

        # Bid/ask spread quality — tight spread = good liquidity
        if c['ask'] > 0 and c['bid'] > 0:
            spread_pct = (c['ask'] - c['bid']) / c['call_premium_mid'] * 100
            if spread_pct < 5:    score += 5
            elif spread_pct > 15: score -= 5

        c['risk_score'] = score
        c['tv_pct'] = round(tv_pct, 2)

        # Risk label
        if score >= 50:    c['risk_label'] = 'LOW RISK'
        elif score >= 30:  c['risk_label'] = 'MODERATE RISK'
        elif score >= 10:  c['risk_label'] = 'ELEVATED RISK'
        else:              c['risk_label'] = 'HIGH RISK'

    # Remove any earnings disqualifiers
    candidates = [c for c in candidates if c.get('earnings_days') is None or c['earnings_days'] > 5]

    # Sort by risk-adjusted score descending
    candidates.sort(key=lambda x: x['risk_score'], reverse=True)

    # Add capital sizing — deploy available cash across top candidates
    try:
        wb = openpyxl.load_workbook(POSITIONS_FILE)
        cfg_total = None
        # Try to get available capital from config
        with open('C:\\TradingBot\\config.json') as f:
            import json as _json
            _cfg = _json.load(f)
            total = _cfg['portfolio']['total_value']
            reserve = _cfg['portfolio']['dry_powder_reserve']
            available = total - reserve
        # Distribute available capital across top 5 candidates evenly
        top_n = min(5, len(candidates))
        if top_n > 0:
            per_position = int(available / top_n / 100) * 100  # round to nearest $100
            for c in candidates[:top_n]:
                shares = int(per_position / c['current'] / 100) * 100  # round to nearest 100 shares
                shares = max(100, shares)  # minimum 1 contract
                c['suggested_shares'] = shares
                c['suggested_capital'] = round(shares * c['current'], 2)
                c['suggested_premium'] = round(shares * c['call_premium_mid'], 2)
                c['suggested_max_gain'] = round(shares * c['net_per_share'], 2)
    except:
        for c in candidates:
            c['suggested_shares'] = 100
            c['suggested_capital'] = round(100 * c['current'], 2)
            c['suggested_premium'] = round(100 * c['call_premium_mid'], 2)
            c['suggested_max_gain'] = round(100 * c['net_per_share'], 2)

    return candidates[:5]

def build_buy_write_section(candidates, macro_events):
    L = []
    L.append('SECTION I - BUY-WRITE CANDIDATES')
    L.append('-' * 40)
    L.append('Buy stock + immediately write ITM call. Ranked by risk-adjusted return.')
    L.append('Scans ITM strikes 0.5%-8% below price. Only shows candidates meeting 1%+ weekly target.')
    L.append('Capital sizing shown assumes equal deployment across top candidates.')
    L.append('')
    if not candidates:
        L.append('No buy-write candidates meeting 1% weekly target today.')
        L.append('IV may be too low, bid/ask spreads too wide, or earnings risk too high.')
        return '\n'.join(L)
    today = date.today()
    next_7_high = [e for e in macro_events if e['impact'] == 'HIGH' and (e['date'] - today).days <= 7]
    macro_warn = next_7_high[0]['name'] if next_7_high else None

    # Check available capital from config
    available_capital = 313000  # default
    expiring_capital = 0
    expiring_tickers = []
    try:
        import json as _json
        with open('C:\\TradingBot\\config.json') as _f:
            _cfg = _json.load(_f)
        _total = _cfg['portfolio']['total_value']
        _reserve = _cfg['portfolio']['dry_powder_reserve']
        available_capital = _total - _reserve
    except:
        pass

    # Check buy-write positions for expiry
    try:
        _wb = openpyxl.load_workbook(POSITIONS_FILE, read_only=True, data_only=True)
        _ws_bw = _wb['BuyWrites']
        _bw_hdrs = [clean_header(c.value) for c in _ws_bw[4]]
        for _row in _ws_bw.iter_rows(min_row=5, values_only=True):
            if _row[0] is None: continue
            if str(_row[0]).startswith('Total') or str(_row[0]).startswith('  '): break
            _bw = {_bw_hdrs[i]: _row[i] for i, h in enumerate(_bw_hdrs) if h}
            _ticker = _bw.get('Ticker','')
            _price = _bw.get('PurchasePrice') or 0
            _shares = _bw.get('Shares') or 0
            _exp = _bw.get('CallExpiry')
            if _exp and _ticker and not isinstance(_price, str):
                _exp_date = _exp.date() if hasattr(_exp, 'date') else _exp
                _days = (_exp_date - date.today()).days
                _cap = float(_price) * int(_shares)
                if _days <= 7:
                    expiring_capital += _cap
                    expiring_tickers.append((_ticker, _days, round(_cap, 0)))
    except:
        pass

    capital_insufficient = available_capital < (candidates[0]['suggested_capital'] if candidates else 0)

    # Summary header
    total_suggested_capital = sum(c.get('suggested_capital', 0) for c in candidates)
    total_suggested_gain = sum(c.get('suggested_max_gain', 0) for c in candidates)
    L.append('PORTFOLIO SUMMARY FOR TODAY')
    L.append('  Candidates found:        ' + str(len(candidates)))
    L.append('  Total capital to deploy: $' + format(int(total_suggested_capital), ','))
    if capital_insufficient and expiring_capital > 0:
        L.append('')
        L.append('*** CAPITAL STATUS: INSUFFICIENT FOR NEW POSITIONS TODAY ***')
        L.append('  Capital tied up in buy-writes expiring this week:')
        for t_ticker, t_days, t_cap in expiring_tickers:
            label = 'THIS FRIDAY' if t_days <= 4 else 'in ' + str(t_days) + ' days'
            L.append('    ' + t_ticker + ' expires ' + label + ' — frees $' + format(int(t_cap), ','))
        L.append('  Total freeing up: $' + format(int(expiring_capital), ','))
        L.append('  RECOMMENDATION: Wait until current positions expire before entering new buy-writes.')
        L.append('  Re-evaluate candidates Monday morning with fresh capital.')
        L.append('')
    elif capital_insufficient:
        L.append('*** WARNING: Insufficient capital for suggested position sizes ***')
        L.append('  Consider reducing share count or waiting for capital to free up.')
        L.append('')
    L.append('  Total max gain if all called: $' + format(int(total_suggested_gain), ','))
    if total_suggested_capital > 0:
        overall_pct = round(total_suggested_gain / total_suggested_capital * 100, 2)
        L.append('  Portfolio net return:    ' + str(overall_pct) + '%')
    if macro_warn:
        L.append('  *** MACRO WARNING: ' + macro_warn + ' within 7 days — consider reducing size ***')
    L.append('')
    L.append('=' * 50)
    L.append('')

    for rank, c in enumerate(candidates, 1):
        analyst = c.get('analyst') or {}
        tech = c.get('tech') or {}
        risk_label = c.get('risk_label', 'MODERATE RISK')
        risk_score = c.get('risk_score', 0)

        # Verdict
        if c['net_pct'] >= 2.0 and 'LOW' in risk_label:
            verdict = 'WRITE NOW'
        elif c['net_pct'] >= 1.5 and 'HIGH RISK' not in risk_label:
            verdict = 'WRITE NOW'
        elif c['net_pct'] >= 1.0 and 'HIGH RISK' not in risk_label:
            verdict = 'CONSIDER'
        else:
            verdict = 'CAUTION'

        # Override verdict if capital is tied up
        if capital_insufficient and expiring_capital > 0:
            if verdict in ['WRITE NOW', 'CONSIDER']:
                verdict = 'WAIT — CAPITAL TIED UP UNTIL FRIDAY'
        elif capital_insufficient:
            if verdict in ['WRITE NOW', 'CONSIDER']:
                verdict = 'WAIT — INSUFFICIENT CAPITAL'

        if macro_warn:
            if verdict == 'WRITE NOW': verdict = 'CONSIDER'

        reasons_for = []
        reasons_against = []

        if c['net_pct'] >= 2.0:
            reasons_for.append('Net return ' + str(c['net_pct']) + '% — well above 1% weekly target')
        else:
            reasons_for.append('Net return ' + str(c['net_pct']) + '% — meets weekly target')
        if c.get('tv_pct', 0) >= 0.5:
            reasons_for.append('Time value $' + str(c['time_value']) + '/share (' + str(c.get('tv_pct',0)) + '%) — cushion above intrinsic')
        rsi = tech.get('rsi', 50)
        bb = tech.get('bb_position', 0.5)
        if rsi < 40:
            reasons_for.append('RSI ' + str(rsi) + ' — oversold, downside support')
        elif rsi > 70:
            reasons_against.append('RSI ' + str(rsi) + ' — overbought, pullback risk')
        if bb < 0.30:
            reasons_for.append('Near lower Bollinger Band — technical support')
        elif bb > 0.80:
            reasons_against.append('Near upper Bollinger Band — stretched')
        beta_risk = analyst.get('beta_risk', 'MODERATE')
        if beta_risk in ['HIGH', 'VERY HIGH']:
            reasons_against.append('Beta ' + str(analyst.get('beta','?')) + ' — ' + beta_risk + ', size conservatively')
        elif beta_risk == 'LOW':
            reasons_for.append('Beta ' + str(analyst.get('beta','?')) + ' — defensive, lower gap risk')
        if analyst.get('rec') in ['STRONG BUY', 'BUY']:
            reasons_for.append('Analyst consensus: ' + analyst['rec'] + ' (' + str(analyst.get('n_analysts','?')) + ' analysts)')
        if analyst.get('upside') and analyst['upside'] > 20:
            reasons_against.append('Analyst target ' + str(analyst['upside']) + '% above current — you are selling upside')
        if c.get('earnings_days') is not None and c['earnings_days'] <= 14:
            reasons_against.append('Earnings in ' + str(c['earnings_days']) + ' days — gap risk')
        if macro_warn:
            reasons_against.append('Macro event within 7 days: ' + macro_warn)

        L.append('#' + str(rank) + ' — ' + c['ticker'] + ' | ' + verdict + ' | ' + risk_label + ' (Score: ' + str(risk_score) + ')')
        L.append('Current price:          $' + str(c['current']) + ' | 52W high: $' + str(c['week_high']) + ' (' + str(c['proximity_pct']) + '% below)')
        L.append('Best call strike:        $' + str(c['call_strike']) + ' (' + str(c['strike_discount_pct']) + '% below price) | Expiry: ' + c['expiry'])
        L.append('Call bid/ask/mid:        $' + str(c['bid']) + ' / $' + str(c['ask']) + ' / $' + str(c['call_premium_mid']) + '/share')
        L.append('  Intrinsic / Time val:  $' + str(c['intrinsic']) + ' / $' + str(c['time_value']) + ' per share')
        L.append('Stock loss if called:    $' + str(abs(c['stock_loss_per_share'])) + '/share')
        L.append('NET RETURN if called:    $' + str(c['net_per_share']) + '/share = ' + str(c['net_pct']) + '%')
        L.append('IV (implied vol):        ' + str(c['iv']) + '%')
        if analyst.get('target'):
            L.append('Analyst:                 $' + str(analyst['target']) + ' target | ' + str(analyst.get('upside','?')) + '% upside | ' + str(analyst.get('rec','?')))
        if tech.get('rsi'):
            L.append('Technicals:              RSI ' + str(tech['rsi']) + ' | BB ' + str(bb) + ' | vs SMA20: ' + str(tech.get('pct_from_sma20','?')) + '%')
        if c.get('earnings_days') is not None:
            L.append('Earnings:                ' + str(c.get('earnings_dt','?')) + ' (' + str(c['earnings_days']) + ' days away)')
        L.append('')
        L.append('RECOMMENDATION: *** ' + verdict + ' ***')
        for r in reasons_for:  L.append('  + ' + r)
        for r in reasons_against: L.append('  - ' + r)
        L.append('')
        # Suggested sizing
        shares = c.get('suggested_shares', 100)
        cap = c.get('suggested_capital', round(c['current']*100, 2))
        prem = c.get('suggested_premium', round(c['call_premium_mid']*100, 2))
        gain = c.get('suggested_max_gain', round(c['net_per_share']*100, 2))
        contracts = int(shares / 100)
        L.append('SUGGESTED SIZING (' + str(shares) + ' shares / ' + str(contracts) + ' contract' + ('s' if contracts > 1 else '') + ')')
        L.append('  Step 1: BUY ' + str(shares) + ' shares ' + c['ticker'])
        L.append('          Cost: $' + str(c['current']) + '/share = $' + format(int(cap), ',') + ' total')
        L.append('  Step 2: SELL TO OPEN CALL | Strike: $' + str(c['call_strike']) + ' | Expiry: ' + c['expiry'])
        L.append('          Contracts: ' + str(contracts) + ' | LIMIT at $' + str(c['call_premium_mid']) + '/share')
        L.append('          Premium collected: $' + format(int(prem), ','))
        L.append('  Net cost basis: $' + str(round(c['current'] - c['call_premium_mid'], 2)) + '/share')
        L.append('  Max gain if called: $' + format(int(gain), ',') + ' (' + str(c['net_pct']) + '%)')
        L.append('*** Verify actual bid/ask in ATP before placing order ***')
        L.append('')
        L.append('-' * 50)
    return '\n'.join(L)

def update_buy_writes_macro_flags(macro_events):
    try:
        wb = openpyxl.load_workbook(POSITIONS_FILE)
        if 'BuyWrites' not in wb.sheetnames:
            return
        ws_bw = wb['BuyWrites']
        for row in range(5, 1005):
            ticker = ws_bw.cell(row=row, column=1).value
            if ticker is None:
                break
            expiry_val = ws_bw.cell(row=row, column=6).value
            if expiry_val:
                flag = get_macro_flag_string(expiry_val, macro_events)
                c = ws_bw.cell(row=row, column=15)
                c.value = flag
                from openpyxl.styles import Font as OFont
                c.font = OFont(color='FFB347' if flag and flag != 'None' else '00E5CC',
                               size=9, name='Calibri')
        safe_save(wb)
        print('  BuyWrites MacroRiskFlag updated')
    except Exception as ex:
        print('Warning: Could not update BuyWrites macro flags: ' + str(ex))

def build_macro_section(open_puts, macro_events):
    L = []
    L.append('SECTION H - MACRO ECONOMIC CALENDAR')
    L.append('-' * 40)
    today = date.today()
    max_expiry = today + timedelta(days=30)
    for p in open_puts:
        try:
            exp = p.get('Expiry')
            if exp:
                exp_date = exp.date() if hasattr(exp, 'date') else exp
                if exp_date > max_expiry:
                    max_expiry = exp_date
        except:
            pass
    L.append('Scanning through ' + max_expiry.strftime('%B %d, %Y') + ' (' + str((max_expiry - today).days) + ' days)')
    L.append('')
    high = [e for e in macro_events if e['impact'] == 'HIGH']
    medium = [e for e in macro_events if e['impact'] == 'MEDIUM']
    low = [e for e in macro_events if e['impact'] == 'LOW']
    if high:
        L.append('*** HIGH IMPACT EVENTS ***')
        for e in high:
            days_away = (e['date'] - today).days
            label = 'TODAY' if days_away == 0 else 'TOMORROW' if days_away == 1 else str(days_away) + ' days away'
            L.append(e['date'].strftime('%a %b %d') + ' (' + label + ') — ' + e['name'])
            L.append('  ' + e['note'])
            affected = []
            for p in open_puts:
                try:
                    exp = p.get('Expiry')
                    if exp:
                        exp_date = exp.date() if hasattr(exp, 'date') else exp
                        if exp_date >= e['date']:
                            affected.append(p['Ticker'] + ' exp ' + exp_date.strftime('%b %d'))
                except:
                    pass
            if affected:
                L.append('  Affects: ' + ', '.join(affected))
            L.append('')
    if medium:
        L.append('MEDIUM IMPACT EVENTS')
        for e in medium:
            L.append(e['date'].strftime('%a %b %d') + ' (' + str((e['date'] - today).days) + ' days) — ' + e['name'])
        L.append('')
    if low:
        L.append('LOW IMPACT (within 10 days)')
        for e in low:
            L.append(e['date'].strftime('%a %b %d') + ' (' + str((e['date'] - today).days) + ' days) — ' + e['name'])
        L.append('')
    next_7 = [e for e in high if (e['date'] - today).days <= 7]
    if next_7:
        L.append('*** CAUTION: ' + str(len(next_7)) + ' HIGH IMPACT event(s) within 7 days ***')
        L.append('Consider reducing new position size or waiting until after announcements.')
    else:
        L.append('No high impact events in next 7 days — clear to write new positions.')
    return '\n'.join(L)

def get_performance_summary():
    try:
        wb = openpyxl.load_workbook(POSITIONS_FILE, read_only=True, data_only=True)
        ws = wb['ClosedTrades']
        headers = [clean_header(c.value) for c in ws[3]]
        trades = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] is None: continue
            trade = {h: row[i] for i, h in enumerate(headers) if h}
            trades.append(trade)
        if not trades: return None
        total = len(trades)
        expired = sum(1 for t in trades if t.get('Outcome') == 'EXPIRED')
        assigned = sum(1 for t in trades if t.get('Outcome') == 'ASSIGNED')
        called_away = sum(1 for t in trades if t.get('Outcome') == 'CALLED_AWAY')
        stop_losses = sum(1 for t in trades if t.get('Outcome') == 'STOP_LOSS')
        btc_count = sum(1 for t in trades if t.get('Outcome') == 'BTC')
        results = []
        for t in trades:
            try:
                prem = float(t.get('PremiumCollected') or 0)
                btc_amt = float(t.get('BTCAmount') or 0)
                call = float(t.get('CallPremium') or 0)
                total_income = (prem - btc_amt) + call
                strike = float(t.get('Strike') or 0)
                contracts = int(t.get('Contracts') or 1)
                days_val = t.get('DaysHeld')
                try: days = int(float(str(days_val)))
                except:
                    try:
                        od = t.get('OpenDate'); cd = t.get('CloseDate')
                        days = (cd - od).days if od and cd else 0
                    except: days = 0
                notional = strike * contracts * 100
                ret_pct = total_income / notional if notional > 0 else 0
                results.append({'total_income':total_income,'ret_pct':ret_pct,'prem':prem,
                                 'btc_amt':btc_amt,'call':call,'days':days})
            except: pass
        if not results: return None
        incomes = [r['total_income'] for r in results]
        wins = sum(1 for i in incomes if i > 0)
        returns = [r['ret_pct'] for r in results if r['ret_pct'] != 0]
        return {
            'total':total,'expired':expired,'assigned':assigned,'called_away':called_away,
            'stop_losses':stop_losses,'btc':btc_count,
            'win_rate':round(wins/total*100,1) if total > 0 else 0,
            'avg_return':round(sum(returns)/len(returns)*100,2) if returns else 0,
            'best':round(max(returns)*100,2) if returns else 0,
            'worst':round(min(returns)*100,2) if returns else 0,
            'total_premium':round(sum(r['prem'] for r in results),2),
            'total_btc_paid':round(sum(r['btc_amt'] for r in results),2),
            'total_call':round(sum(r['call'] for r in results),2),
            'total_income':round(sum(r['total_income'] for r in results),2),
            'avg_days':round(sum(r['days'] for r in results if r['days']>0)/max(1,sum(1 for r in results if r['days']>0)),1)
        }
    except Exception as ex:
        print('Warning: Could not read performance data: ' + str(ex))
        return None

def build_performance_section():
    L = []
    L.append('SECTION F - PERFORMANCE SUMMARY')
    L.append('-' * 40)
    perf = get_performance_summary()
    if perf is None:
        L.append('No closed trades on file yet.')
    else:
        L.append('Total trades:             ' + str(perf['total']))
        L.append('Puts expired worthless:   ' + str(perf['expired']))
        L.append('Puts assigned:            ' + str(perf['assigned']))
        L.append('Covered calls closed:     ' + str(perf['called_away']))
        L.append('Bought to close:          ' + str(perf['btc']))
        L.append('Stop losses:              ' + str(perf['stop_losses']))
        L.append('Win rate:                 ' + str(perf['win_rate']) + '%')
        L.append('Avg return per trade:     ' + str(perf['avg_return']) + '%')
        L.append('Best trade:               ' + str(perf['best']) + '%')
        L.append('Worst trade:              ' + str(perf['worst']) + '%')
        L.append('Total premium collected:  $' + format(perf['total_premium'], ','))
        L.append('Total BTC paid:           $' + format(perf['total_btc_paid'], ','))
        L.append('Total call premium:       $' + format(perf['total_call'], ','))
        L.append('Total net income:         $' + format(perf['total_income'], ','))
        L.append('Avg days held:            ' + str(perf['avg_days']))
    L.append('')
    L.append('=' * 60)
    L.append('IMPORTANT: All recommendations are estimates only.')
    L.append('Verify all premiums and strikes in Fidelity ATP')
    L.append('before placing any order. This is not financial advice.')
    L.append('=' * 60)
    return '\n'.join(L)

def send_email(cfg, subject, body):
    e = cfg['email']
    msg = MIMEMultipart()
    msg['From'] = e['sender']
    msg['To'] = e['recipient']
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=ctx) as server:
        server.login(e['sender'], e['app_password'])
        server.sendmail(e['sender'], e['recipient'], msg.as_string())

def main():
    print('Loading config...')
    cfg = load_config()
    rules = cfg['rules']
    print('Checking market calendar...')
    if not is_market_open_today():
        print('Market is closed today. No report generated.')
        return
    print('Loading positions...')
    open_puts, assigned, buy_writes = load_positions()
    print('Positions loaded: ' + str(len(open_puts)) + ' puts, ' + str(len(assigned)) + ' assigned, ' + str(len(buy_writes)) + ' buy-writes')
    if buy_writes:
        bw_cap = sum(float(bw.get('PurchasePrice',0)) * int(bw.get('Shares',0))
                     for bw in buy_writes
                     if not isinstance(bw.get('PurchasePrice'), str) and bw.get('PurchasePrice'))
        print('Buy-write capital: $' + format(int(bw_cap), ','))
    open_position_tickers = ([p['Ticker'] for p in open_puts] +
                             [a['Ticker'] for a in assigned] +
                             [bw['Ticker'] for bw in buy_writes if bw.get('Ticker')])
    watchlist = cfg['watchlist']
    print('Fetching VIX and SPY...')
    vix = get_vix()
    spy_trend = get_spy_trend()
    print('VIX: ' + str(vix))
    print('Building macro calendar...')
    today = date.today()
    max_expiry = today + timedelta(days=30)
    for p in open_puts:
        try:
            exp = p.get('Expiry')
            if exp:
                exp_date = exp.date() if hasattr(exp, 'date') else exp
                if exp_date > max_expiry:
                    max_expiry = exp_date
        except:
            pass
    macro_events = get_macro_events(today, max_expiry)
    print('Fetching stock data for ' + str(len(watchlist)) + ' stocks...')
    all_data = {}
    for ticker in watchlist:
        print('  Fetching ' + ticker + '...')
        all_data[ticker] = get_stock_data(ticker)
    print('Checking earnings...')
    earnings_tickers = get_earnings_tickers(watchlist)
    print('Identifying watchlist flags...')
    watchlist_flags = [{'ticker': t, 'change': d['pre_market_change_pct']}
                       for t, d in all_data.items() if d and abs(d['pre_market_change_pct']) >= 5.0]
    print('Calculating technicals for all watchlist stocks...')
    tech_cache = {}
    for ticker in watchlist:
        print('  Technicals: ' + ticker + '...')
        tech_cache[ticker] = calculate_technicals(ticker)
    print('Finding sympathy drop candidates...')
    candidates = find_put_candidates(cfg, all_data, earnings_tickers, open_position_tickers)
    longer_candidates = find_longer_dated_candidates(cfg, all_data, earnings_tickers, open_position_tickers)
    print('Running mean reversion screen...')
    mean_reversion_candidates = find_mean_reversion_candidates(cfg, all_data, earnings_tickers, open_position_tickers, tech_cache)
    print('Finding buy-write candidates...')
    buy_write_candidates = find_buy_write_candidates(cfg, all_data, earnings_tickers, tech_cache)
    print('Updating Excel file (MacroRiskFlag, TechnicalSnapshot, MacroCalendar, BuyWrites)...')
    update_excel(open_puts, macro_events, tech_cache)
    update_buy_writes_macro_flags(macro_events)
    print('Reloading open puts and assigned after Excel update...')
    open_puts_fresh, assigned_fresh, _ = load_positions()
    # Keep buy_writes from original load — Excel formulas corrupt on reload
    open_puts = open_puts_fresh
    assigned = assigned_fresh
    print('Building report...')
    regime_section = build_market_regime_section(vix, spy_trend, macro_events)
    report = build_report(cfg, candidates, earnings_tickers, all_data, watchlist_flags, open_puts, assigned, buy_writes, macro_events, vix)
    longer_section = build_longer_dated_section(longer_candidates, macro_events)
    mean_reversion_section = build_mean_reversion_section(mean_reversion_candidates, macro_events)
    sections_cd = build_sections_cd(open_puts, assigned, all_data, rules)
    perf_section = build_performance_section()
    macro_section = build_macro_section(open_puts, macro_events)
    buy_write_section = build_buy_write_section(buy_write_candidates, macro_events)
    full_report = (regime_section + '\n' + report + '\n' + longer_section + '\n' +
                   mean_reversion_section + '\n' + buy_write_section + '\n' +
                   sections_cd + '\n' + perf_section + '\n' + macro_section)
    print('')
    print(full_report)
    print('')
    print('Sending email...')
    send_email(cfg, 'Trading Bot Morning Report - ' + datetime.now().strftime('%Y-%m-%d'), full_report)
    print('Report emailed successfully.')

if __name__ == '__main__':
    main()
